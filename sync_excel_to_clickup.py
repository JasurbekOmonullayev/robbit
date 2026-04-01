import argparse
import json
import os
import re
import sys
import time
from decimal import Decimal, InvalidOperation
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from urllib import error, parse, request

import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build

DEFAULT_EXCEL = "Jiddi_lead.xlsx"
DEFAULT_LEADS_SHEET = "Leads"
DEFAULT_MAPPING_SHEET = "columns"
LEAD_COLUMN_NAME = "2. Lead ID"
TASK_NAME_COLUMN = "3. Ismingizni kiriting!"
MAPPING_SOURCE_HEADER = "sheets_columns_name"
MAPPING_CLICKUP_ID_HEADER = "clickup_column_id"
DEFAULT_STATE_FILE = ".sync_state.json"


def load_env_file(env_path: Path) -> None:
    if not env_path.exists():
        return
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip().lstrip("\ufeff")
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def load_local_state_ids(path: Path) -> set[str]:
    if not path.exists():
        return set()
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:  # noqa: BLE001
        return set()
    ids = data.get("synced_lead_ids", [])
    if not isinstance(ids, list):
        return set()
    return {normalize_text(x) for x in ids if normalize_text(x)}


def save_local_state_ids(path: Path, ids: set[str]) -> None:
    payload = {"synced_lead_ids": sorted(ids)}
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


class ClickUpClient:
    def __init__(self, token: str, timeout: int = 30, retries: int = 3) -> None:
        self.token = token
        self.timeout = timeout
        self.retries = retries

    def _request(self, method: str, url: str, payload: dict[str, Any] | None = None) -> dict[str, Any]:
        body = None
        headers = {
            "Authorization": self.token,
            "Content-Type": "application/json",
        }
        if payload is not None:
            body = json.dumps(payload).encode("utf-8")

        last_error: Exception | None = None
        for attempt in range(1, self.retries + 1):
            req = request.Request(url=url, data=body, headers=headers, method=method)
            try:
                with request.urlopen(req, timeout=self.timeout) as resp:
                    text = resp.read().decode("utf-8")
                    if not text:
                        return {}
                    return json.loads(text)
            except error.HTTPError as exc:
                detail = exc.read().decode("utf-8", errors="ignore")
                last_error = RuntimeError(f"HTTP {exc.code} for {method} {url}: {detail}")
                if exc.code in {429, 500, 502, 503, 504} and attempt < self.retries:
                    time.sleep(attempt)
                    continue
                raise last_error
            except Exception as exc:  # noqa: BLE001
                last_error = exc
                if attempt < self.retries:
                    time.sleep(attempt)
                    continue
                raise

        raise RuntimeError(f"Request failed: {last_error}")

    def list_tasks(self, list_id: str) -> list[dict[str, Any]]:
        tasks: list[dict[str, Any]] = []
        page = 0
        while True:
            qs = parse.urlencode({"include_closed": "true", "page": page})
            url = f"https://api.clickup.com/api/v2/list/{list_id}/task?{qs}"
            data = self._request("GET", url)
            chunk = data.get("tasks", [])
            if not chunk:
                break
            tasks.extend(chunk)
            page += 1
        return tasks

    def list_fields(self, list_id: str) -> dict[str, dict[str, Any]]:
        url = f"https://api.clickup.com/api/v2/list/{list_id}/field"
        data = self._request("GET", url)
        fields: dict[str, dict[str, Any]] = {}
        for item in data.get("fields", []):
            fid = normalize_text(item.get("id"))
            if fid:
                fields[fid] = item
        return fields

    def create_task(self, list_id: str, payload: dict[str, Any]) -> dict[str, Any]:
        url = f"https://api.clickup.com/api/v2/list/{list_id}/task"
        return self._request("POST", url, payload)


class GoogleSheetsReader:
    def __init__(self, credentials_file: str, spreadsheet_id: str) -> None:
        creds = service_account.Credentials.from_service_account_file(
            credentials_file,
            scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"],
        )
        self.service = build("sheets", "v4", credentials=creds)
        self.spreadsheet_id = spreadsheet_id

    def read_sheet(self, sheet_name: str) -> list[list[Any]]:
        resp = (
            self.service.spreadsheets()
            .values()
            .get(
                spreadsheetId=self.spreadsheet_id,
                range=f"{sheet_name}!A:ZZ",
                valueRenderOption="UNFORMATTED_VALUE",
                dateTimeRenderOption="FORMATTED_STRING",
            )
            .execute()
        )
        return resp.get("values", [])


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_choice_key(value: str) -> str:
    return re.sub(r"[\W_]+", "", value.lower(), flags=re.UNICODE)


def is_status_like(text: Any) -> bool:
    key = normalize_choice_key(normalize_text(text))
    return "status" in key or "статус" in key


def to_number(value: Any) -> float | int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value

    text = normalize_text(value)
    if not text:
        return None

    text = text.replace("l:", "").replace("L:", "")
    text = text.replace(",", ".")
    text = re.sub(r"[^0-9eE+.-]", "", text)
    if not text:
        return None

    try:
        number = Decimal(text)
    except InvalidOperation:
        return None

    if number == number.to_integral_value():
        return int(number)
    return float(number)


def canonical_lead_id(value: Any) -> str:
    number = to_number(value)
    if number is None:
        return ""
    return str(number)


def date_to_millis(value: Any) -> int | None:
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, date):
        dt = datetime(value.year, value.month, value.day)
    elif isinstance(value, (int, float)):
        # Google Sheets serial date: 1899-12-30 dan boshlab kunlar
        serial = float(value)
        if 10000 <= serial <= 100000:
            base = datetime(1899, 12, 30)
            dt = base + timedelta(days=serial)
        else:
            return None
    else:
        text = normalize_text(value)
        if not text:
            return None

        patterns = [
            "%Y-%m-%d",
            "%d.%m.%Y",
            "%d/%m/%Y",
            "%m/%d/%Y",
            "%d-%m-%Y",
            "%Y/%m/%d",
        ]
        dt = None
        for pattern in patterns:
            try:
                dt = datetime.strptime(text, pattern)
                break
            except ValueError:
                continue
        if dt is None:
            return None

    return int(dt.timestamp() * 1000)


def extract_spreadsheet_id(value: str) -> str:
    value = normalize_text(value)
    if not value:
        return ""
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", value)
    if m:
        return m.group(1)
    return value


def rows_to_records(rows: list[list[Any]]) -> list[dict[str, Any]]:
    if not rows:
        return []

    headers = [normalize_text(h) for h in rows[0]]
    if not any(headers):
        return []

    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        rec: dict[str, Any] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            rec[h] = row[i] if i < len(row) else ""
        records.append(rec)
    return records


def read_excel_records(excel_path: Path, leads_sheet: str, mapping_sheet: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if leads_sheet not in wb.sheetnames or mapping_sheet not in wb.sheetnames:
        raise RuntimeError(f"'{leads_sheet}' yoki '{mapping_sheet}' sheet topilmadi")

    def sheet_records(sheet_name: str) -> list[dict[str, Any]]:
        ws = wb[sheet_name]
        headers: list[str] = []
        for c in range(1, ws.max_column + 1):
            headers.append(normalize_text(ws.cell(row=1, column=c).value))

        records: list[dict[str, Any]] = []
        for r in range(2, ws.max_row + 1):
            rec: dict[str, Any] = {}
            for idx, h in enumerate(headers, start=1):
                if not h:
                    continue
                rec[h] = ws.cell(row=r, column=idx).value
            records.append(rec)
        return records

    return sheet_records(leads_sheet), sheet_records(mapping_sheet)


def load_mapping(mapping_records: list[dict[str, Any]]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for rec in mapping_records:
        source_name = normalize_text(rec.get(MAPPING_SOURCE_HEADER))
        clickup_id = normalize_text(rec.get(MAPPING_CLICKUP_ID_HEADER))
        if is_status_like(source_name):
            continue
        if source_name and clickup_id:
            mapping[source_name] = clickup_id
    if not mapping:
        raise RuntimeError("Mapping bo'sh: columns sheetdagi clickup_column_id larni tekshiring")
    return mapping


def extract_existing_leads(tasks: list[dict[str, Any]], lead_field_id: str) -> set[str]:
    existing: set[str] = set()
    for task in tasks:
        for field in task.get("custom_fields", []):
            if normalize_text(field.get("id")) != lead_field_id:
                continue
            lead = canonical_lead_id(field.get("value"))
            if lead:
                existing.add(lead)
    return existing


def convert_custom_field_value(raw_value: Any, field_meta: dict[str, Any]) -> Any:
    field_type = normalize_text(field_meta.get("type"))

    if field_type == "date":
        return date_to_millis(raw_value)

    if field_type in {"short_text", "text", "url", "email", "phone"}:
        return normalize_text(raw_value)

    if field_type == "number":
        return to_number(raw_value)

    if field_type == "drop_down":
        raw_text = normalize_text(raw_value)
        if not raw_text:
            return None
        raw_key = normalize_choice_key(raw_text)
        options = (field_meta.get("type_config") or {}).get("options", [])
        for option in options:
            opt_name = normalize_text(option.get("name"))
            if normalize_choice_key(opt_name) == raw_key:
                return option.get("id")
        for option in options:
            opt_key = normalize_choice_key(normalize_text(option.get("name")))
            if raw_key and (raw_key in opt_key or opt_key in raw_key):
                return option.get("id")
        return None

    return normalize_text(raw_value)


def build_rows(
    lead_records: list[dict[str, Any]],
    mapping: dict[str, str],
    field_meta_by_id: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen_in_source: set[str] = set()

    for rec in lead_records:
        lead = canonical_lead_id(rec.get(LEAD_COLUMN_NAME))
        if not lead:
            continue
        if lead in seen_in_source:
            continue
        seen_in_source.add(lead)

        raw_name = normalize_text(rec.get(TASK_NAME_COLUMN))
        task_name = raw_name if raw_name else f"Lead {lead}"

        custom_fields: list[dict[str, Any]] = []
        for sheet_col_name, clickup_field_id in mapping.items():
            raw_value = rec.get(sheet_col_name)
            if raw_value is None or raw_value == "":
                continue

            meta = field_meta_by_id.get(clickup_field_id, {})
            if is_status_like(meta.get("name")) or normalize_text(meta.get("type")) == "status":
                continue
            value = convert_custom_field_value(raw_value, meta)
            if value is None or value == "":
                continue

            custom_fields.append({"id": clickup_field_id, "value": value})

        rows.append({"lead_id": lead, "task_name": task_name, "custom_fields": custom_fields})

    return rows


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Sync Leads to ClickUp")
    parser.add_argument("--source", choices=["google", "excel"], default="google", help="Ma'lumot manbasi")
    parser.add_argument("--excel", default=DEFAULT_EXCEL, help="Excel fayl manzili (source=excel)")
    parser.add_argument("--dry-run", action="store_true", help="Task yaratmaydi, faqat hisobot")
    parser.add_argument("--limit", type=int, default=0, help="Faqat N ta yangi leadni qayta ishlash")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    load_env_file(Path(".env"))

    token = os.getenv("CLICKUP_API_TOKEN", "").strip()
    list_id = os.getenv("CLICKUP_LIST_ID", "").strip()
    lead_field_id = os.getenv("CLICKUP_LEAD_ID_FIELD_ID", "").strip()

    leads_sheet_name = os.getenv("GOOGLE_LEADS_SHEET_NAME", DEFAULT_LEADS_SHEET).strip() or DEFAULT_LEADS_SHEET
    mapping_sheet_name = os.getenv("GOOGLE_MAPPING_SHEET_NAME", DEFAULT_MAPPING_SHEET).strip() or DEFAULT_MAPPING_SHEET
    state_file = Path(os.getenv("SYNC_STATE_FILE", DEFAULT_STATE_FILE).strip() or DEFAULT_STATE_FILE)

    if not token:
        print("ERROR: CLICKUP_API_TOKEN topilmadi")
        return 1
    if not list_id:
        print("ERROR: CLICKUP_LIST_ID topilmadi")
        return 1
    if not lead_field_id:
        print("ERROR: CLICKUP_LEAD_ID_FIELD_ID topilmadi")
        return 1

    try:
        if args.source == "google":
            credentials_file = os.getenv("GOOGLE_SERVICE_ACCOUNT_FILE", "").strip()
            spreadsheet_raw = os.getenv("GOOGLE_SPREADSHEET_ID", "").strip() or os.getenv("GOOGLE_SHEETS_URL", "").strip()
            spreadsheet_id = extract_spreadsheet_id(spreadsheet_raw)

            if not credentials_file:
                print("ERROR: GOOGLE_SERVICE_ACCOUNT_FILE topilmadi")
                return 1
            if not spreadsheet_id:
                print("ERROR: GOOGLE_SPREADSHEET_ID yoki GOOGLE_SHEETS_URL topilmadi")
                return 1

            reader = GoogleSheetsReader(credentials_file=credentials_file, spreadsheet_id=spreadsheet_id)
            lead_records = rows_to_records(reader.read_sheet(leads_sheet_name))
            mapping_records = rows_to_records(reader.read_sheet(mapping_sheet_name))
        else:
            excel_path = Path(args.excel)
            if not excel_path.exists():
                print(f"ERROR: Excel fayl topilmadi: {excel_path}")
                return 1
            lead_records, mapping_records = read_excel_records(excel_path, leads_sheet_name, mapping_sheet_name)
    except Exception as exc:  # noqa: BLE001
        print(f"ERROR: manbani o'qishda xatolik: {exc}")
        print("Maslahat: service account email'ni Google Sheetga share qilinganini tekshiring.")
        return 1

    mapping = load_mapping(mapping_records)
    # Force Lead ID mapping to env-configured field id so columns sheetdagi eski id xatoga sabab bo'lmaydi.
    mapping[LEAD_COLUMN_NAME] = lead_field_id

    client = ClickUpClient(token=token)
    try:
        tasks = client.list_tasks(list_id)
        field_meta_by_id = client.list_fields(list_id)
    except Exception as exc:  # noqa: BLE001
        print(f"ERROR: ClickUp task listni o'qib bo'lmadi: {exc}")
        return 1

    rows = build_rows(lead_records, mapping, field_meta_by_id)
    clickup_existing = extract_existing_leads(tasks, lead_field_id)
    local_existing = load_local_state_ids(state_file)
    existing_leads = clickup_existing | local_existing

    created = 0
    skipped = 0
    errors = 0

    process_count = 0
    for row in rows:
        lead_id = row["lead_id"]

        if lead_id in existing_leads:
            skipped += 1
            continue

        if args.limit and process_count >= args.limit:
            break

        payload = {
            "name": row["task_name"],
            "custom_fields": row["custom_fields"],
        }

        if args.dry_run:
            print(f"DRY-RUN create: lead_id={lead_id} name={row['task_name']}")
            created += 1
            existing_leads.add(lead_id)
            process_count += 1
            continue

        try:
            client.create_task(list_id=list_id, payload=payload)
            created += 1
            existing_leads.add(lead_id)
            local_existing.add(lead_id)
            process_count += 1
            print(f"CREATED: lead_id={lead_id}")
        except Exception as exc:  # noqa: BLE001
            errors += 1
            print(f"ERROR create lead_id={lead_id}: {exc}")

    print("--- SUMMARY ---")
    print(f"source_rows_prepared={len(rows)}")
    print(f"existing_in_clickup={len(tasks)} tasks, {len(existing_leads)} unique_lead_ids_after_run")
    print(f"created={created}")
    print(f"skipped_existing={skipped}")
    print(f"errors={errors}")
    if not args.dry_run:
        save_local_state_ids(state_file, local_existing)
    return 0 if errors == 0 else 2


if __name__ == "__main__":
    sys.exit(main())
